# pyright: ignore
import pandas as pd
import glob
import os
import logging
from typing import List, Optional, Dict, Tuple, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from copy import copy as cell_copy
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import configparser

# カスタム例外クラス
class SalesDataError(Exception):
    """売上データ処理の基本例外クラス"""
    pass

class FileNotFoundError(SalesDataError):
    """ファイルが見つからない場合の例外"""
    pass

class DataValidationError(SalesDataError):
    """データ検証エラーの例外"""
    pass

class ExcelOperationError(SalesDataError):
    """Excel操作エラーの例外"""
    pass

# ログ設定
def setup_logging() -> None:
    """ログ設定を初期化"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILE, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

logger = logging.getLogger(__name__)

# エラーハンドリングデコレータ
def safe_excel_operation(func):
    """Excel操作を安全に実行するデコレータ"""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except FileNotFoundError as e:
            logger.error(f"ファイルが見つかりません: {e}")
            raise
        except PermissionError as e:
            logger.error(f"ファイルアクセス権限エラー: {e}")
            raise ExcelOperationError(f"ファイルアクセス権限エラー: {e}")
        except Exception as e:
            logger.error(f"予期しないエラー ({func.__name__}): {e}")
            raise ExcelOperationError(f"予期しないエラー: {e}")
    return wrapper

def safe_data_operation(func):
    """データ処理を安全に実行するデコレータ"""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except pd.errors.EmptyDataError as e:
            logger.error(f"データが空です: {e}")
            raise DataValidationError(f"データが空です: {e}")
        except pd.errors.ParserError as e:
            logger.error(f"データ解析エラー: {e}")
            raise DataValidationError(f"データ解析エラー: {e}")
        except Exception as e:
            logger.error(f"データ処理エラー ({func.__name__}): {e}")
            raise DataValidationError(f"データ処理エラー: {e}")
    return wrapper

# 設定ファイル読み込み
config = configparser.ConfigParser()
config.read('Config.ini', encoding='utf-8')

# 設定値の取得
INPUT_FOLDER: str = config['input']['folder']
OUTPUT_FILE: str = config['output']['file']
MASTER_FILE: str = config['input']['master_file']
MASTER_SHEET: str = config['input']['master_sheet']
TEMPLATE_FILE: str = config['input']['template_file']
SALES_DB_FILE: str = config['input']['sales_db_file']
LOG_FILE: str = config['output']['log_file']

# 列名・ラベル
COL_CUSTOMER_CODE = config['columns']['customer_code']
COL_CUSTOMER_NAME = config['columns']['customer_name']
COL_SALES_AMOUNT = config['columns']['sales_amount']
COL_GROUP_NAME = config['columns']['group_name']
COL_GROUP_LABEL = config['columns']['group_label']
COL_TOTAL_LABEL = config['columns']['total_label']
COL_TOTAL_SALES = config['columns']['total_sales']

# シート名
SHEET_CONCAT = config['sheets']['concat']
SHEET_COMPANY = config['sheets']['company']
SHEET_DISPERSION = config['sheets']['dispersion']
SHEET_GROUP_RANK = config['sheets']['group_rank']

# 色
COLOR_UNKNOWN_GROUP = config['format']['unknown_group_color']
COLOR_GROUP = config['format']['group_color']
COLOR_TOTAL_ROW = config['format']['total_row_color']

# 追加の設定値取得
YELLOW = config['format']['yellow']
LIGHT_GREEN = config['format']['light_green']
LIGHT_YELLOW = config['format']['light_yellow']
UNMATCHED_ROW_COLOR = config['format']['unmatched_row_color']
BORDER_COLOR = config['format']['border_color']
BORDER_STYLE = config['format']['border_style']
SIMILARITY_THRESHOLD = float(config['threshold']['similarity'])
KEI_COL = int(config['dispersion']['kei_col'])
SAGI_COL = int(config['dispersion']['sagi_col'])
SUM_FORMULA = config['dispersion']['sum_formula']
DIFF_FORMULA = config['dispersion']['diff_formula']
LABEL_RANK = config['labels']['rank']
LABEL_CUSTOMER = config['labels']['customer']
LABEL_CURRENT_SALES = config['labels']['current_sales']
LABEL_TOTAL = config['labels']['total']
LABEL_PREV_SALES = config['labels']['prev_sales']
LABEL_LAST_SALES = config['labels']['last_sales']
LABEL_NOW_SALES = config['labels']['now_sales']
TEMPLATE_SHEETS_LIST = [s.strip() for s in config['template']['sheets'].split(',')]
EXCLUDE_AUTO_WIDTH = [s.strip() for s in config['exclude_sheets']['auto_width'].split(',')]

# グローバル変数
YYMM: str = ''
PREV_YYMM: str = ''
LAST_MONTH_YYMM: str = ''
TEMPLATE_SHEETS: List[str] = []
SALES_DB_SHEETS: List[str] = []

def copy_cell(src_cell: Any, dst_cell: Any) -> None:
    """openpyxlのセル内容・スタイルを安全にコピーする"""
    if isinstance(dst_cell, MergedCell) or isinstance(src_cell, MergedCell):
        return
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        if src_cell.font:
            dst_cell.font = cell_copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = cell_copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = cell_copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = cell_copy(src_cell.alignment)
        if src_cell.number_format:
            dst_cell.number_format = src_cell.number_format
        if src_cell.protection:
            dst_cell.protection = cell_copy(src_cell.protection)
    if src_cell.hyperlink:
        dst_cell.hyperlink = src_cell.hyperlink
    if src_cell.comment:
        dst_cell.comment = cell_copy(src_cell.comment)

def get_prev_yymm(yymm: str) -> str:
    yy = int(yymm[:2])
    mm = int(yymm[2:])
    prev_yy = yy - 1
    return f"{prev_yy:02d}{mm:02d}"

def get_last_month_yymm(yymm: str) -> str:
    yy = int(yymm[:2])
    mm = int(yymm[2:])
    if mm == 1:
        return f"{yy-1:02d}12"
    else:
        return f"{yy:02d}{mm-1:02d}"

@safe_data_operation
def load_and_preprocess_data() -> List[pd.DataFrame]:
    """各社売上フォルダからデータを読み込み、合計行除外・K列計算・列抽出を行う"""
    files = glob.glob(os.path.join(INPUT_FOLDER, '*.xlsx'))
    if not files:
        logger.warning(f"入力フォルダ '{INPUT_FOLDER}' にExcelファイルが見つかりません")
        raise FileNotFoundError(f"入力フォルダ '{INPUT_FOLDER}' にExcelファイルが見つかりません")
    
    all_list: List[pd.DataFrame] = []
    for file in files:
        logger.info(f"ファイル処理中: {file}")
        try:
            df = pd.read_excel(file, engine='openpyxl', dtype=str)
            mask = ~df.iloc[:, 0].astype(str).str.contains('合計', na=False) & ~df.iloc[:, 1].astype(str).str.contains('合計', na=False)
            df = df[mask]
            if df.shape[1] >= 11:
                k = pd.to_numeric(df.iloc[:, 10], errors='coerce')
                g = pd.to_numeric(df.iloc[:, 6], errors='coerce')
                df.iloc[:, 10] = k - g
            df = df.iloc[:, [0, 1, 10]]
            df.columns = [COL_CUSTOMER_CODE, COL_CUSTOMER_NAME, COL_SALES_AMOUNT]
            df[COL_SALES_AMOUNT] = pd.to_numeric(df[COL_SALES_AMOUNT], errors='coerce')
            all_list.append(df)
            logger.info(f"ファイル処理完了: {file} (行数: {len(df)})")
        except Exception as e:
            logger.error(f"ファイル処理エラー {file}: {e}")
            raise
    
    if not all_list:
        raise DataValidationError("有効なデータが見つかりませんでした")
    
    logger.info(f"全ファイル処理完了: {len(all_list)}ファイル")
    return all_list

@safe_data_operation
def merge_with_master(all_df: pd.DataFrame) -> pd.DataFrame:
    """変換マスタとマージし、列順・型・ソートを整える"""
    if not os.path.exists(MASTER_FILE):
        logger.error(f"マスタファイルが見つかりません: {MASTER_FILE}")
        raise FileNotFoundError(f"マスタファイルが見つかりません: {MASTER_FILE}")
    
    try:
        master_df = pd.read_excel(MASTER_FILE, sheet_name=MASTER_SHEET, dtype=str)
        logger.info(f"マスタファイル読み込み完了: {len(master_df)}行")
    except Exception as e:
        logger.error(f"マスタファイル読み込みエラー: {e}")
        raise
    
    if len(master_df.columns) < 2:
        raise DataValidationError("マスタファイルの列数が不足しています（最低2列必要）")
    
    master_df = master_df.iloc[:, :2]
    master_df.columns = [COL_GROUP_NAME, COL_CUSTOMER_NAME]
    # 全角・半角空白削除で結合精度UP
    all_df[COL_CUSTOMER_NAME] = all_df[COL_CUSTOMER_NAME].str.replace(' ', '').str.replace('　', '')
    master_df[COL_CUSTOMER_NAME] = master_df[COL_CUSTOMER_NAME].str.replace(' ', '').str.replace('　', '')
    
    merged_df = all_df.merge(master_df, how='left', on=COL_CUSTOMER_NAME)
    merged_df = merged_df[[COL_CUSTOMER_CODE, COL_CUSTOMER_NAME, COL_SALES_AMOUNT, COL_GROUP_NAME]]
    merged_df[COL_SALES_AMOUNT] = pd.to_numeric(merged_df[COL_SALES_AMOUNT], errors='coerce')
    merged_df = merged_df.sort_values(by=COL_SALES_AMOUNT, ascending=False)
    
    # グループ未設定のデータ数をチェック
    unknown_groups = merged_df[COL_GROUP_NAME].isna().sum()
    if unknown_groups > 0:
        logger.warning(f"グループ未設定のデータ: {unknown_groups}件")
    
    logger.info(f"マスタマージ完了: {len(merged_df)}行")
    return merged_df

def save_concat_sheet(all_df: pd.DataFrame) -> None:
    """連結シートを出力し、列幅を自動調整"""
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    all_df['今回請求額'] = pd.to_numeric(all_df['今回請求額'], errors='coerce')
    all_df.to_excel(OUTPUT_FILE, index=False, sheet_name=SHEET_CONCAT)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_CONCAT]
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ''
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[col_letter].width = max_length + 2
    format_concat_sheet_colors(ws, all_df)
    wb.save(OUTPUT_FILE)

def format_concat_sheet_colors(ws: Worksheet, all_df: pd.DataFrame) -> None:
    """連結シートでグループ名が不明な行を薄い赤色で色付け"""
    red_fill = PatternFill(start_color=COLOR_UNKNOWN_GROUP, end_color=COLOR_UNKNOWN_GROUP, fill_type='solid')
    for i, (idx, row) in enumerate(all_df.iterrows(), 2):
        group_value = row[COL_GROUP_NAME]
        is_unknown = False
        if pd.isna(group_value):
            is_unknown = True
        elif group_value is None:
            is_unknown = True
        elif str(group_value).strip() == '' or str(group_value).strip() in ['未設定', 'nan', 'none', 'null', 'NaN', 'None', 'NULL', '－', '-']:
            is_unknown = True
        elif str(group_value).replace(' ', '').replace('　', '') == '':
            is_unknown = True
        if is_unknown:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=i, column=col).fill = red_fill

def save_group_summary(all_df: pd.DataFrame) -> pd.DataFrame:
    """グループごとの合計売上を当月売上シートに出力"""
    group_sum = all_df.groupby(COL_GROUP_NAME, dropna=False)[COL_SALES_AMOUNT].sum().reset_index()
    group_sum = group_sum.rename(columns={COL_GROUP_NAME: COL_GROUP_LABEL, COL_SALES_AMOUNT: COL_TOTAL_SALES})
    group_sum = group_sum.sort_values(COL_TOTAL_SALES, ascending=False)
    total = group_sum[COL_TOTAL_SALES].sum()
    total_row = pd.DataFrame({COL_GROUP_LABEL: [COL_TOTAL_LABEL], COL_TOTAL_SALES: [total]})
    group_sum = pd.concat([group_sum, total_row], ignore_index=True)
    group_sum[COL_TOTAL_SALES] = pd.to_numeric(group_sum[COL_TOTAL_SALES], errors='coerce')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        group_sum.to_excel(writer, index=False, sheet_name=YYMM)
    return group_sum

def save_company_sheet(all_df: pd.DataFrame, group_sum: pd.DataFrame) -> None:
    """会社別シートをグループ売上順で出力"""
    group_order = group_sum[group_sum[COL_GROUP_LABEL] != COL_TOTAL_LABEL][COL_GROUP_LABEL].tolist()
    group_rank = {g: i for i, g in enumerate(group_order)}
    all_df['グループ売上順位'] = all_df[COL_GROUP_NAME].map(group_rank).fillna(len(group_order))
    all_df_sorted = all_df.sort_values(['グループ売上順位', COL_SALES_AMOUNT], ascending=[True, False])
    all_df_sorted[COL_SALES_AMOUNT] = pd.to_numeric(all_df_sorted[COL_SALES_AMOUNT], errors='coerce')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        all_df_sorted.drop(columns=['グループ売上順位']).to_excel(writer, index=False, sheet_name=SHEET_COMPANY)

def copy_template_sheets() -> None:
    """テンプレートファイルから指定シートを完全コピー"""
    template_wb = load_workbook(TEMPLATE_FILE)
    output_wb = load_workbook(OUTPUT_FILE)
    for sheet_name in TEMPLATE_SHEETS:
        base_name = sheet_name[4:]
        if base_name not in template_wb.sheetnames:
            continue
        if sheet_name in output_wb.sheetnames:
            del output_wb[sheet_name]
        src_ws = template_wb[base_name]
        ws = output_wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
        for merged_cell_range in src_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_cell_range))
        for row_idx, row_dim in src_ws.row_dimensions.items():
            ws.row_dimensions[row_idx].height = row_dim.height
        for col, col_dim in src_ws.column_dimensions.items():
            ws.column_dimensions[col].width = col_dim.width
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                copy_cell(cell, new_cell)
                if cell.data_type == 'f':
                    new_cell.value = f'={cell.value}' if not str(cell.value).startswith('=') else cell.value
        if src_ws.auto_filter.ref:
            ws.auto_filter.ref = src_ws.auto_filter.ref
        if src_ws.freeze_panes:
            ws.freeze_panes = src_ws.freeze_panes
        ws.page_setup = src_ws.page_setup
        ws.page_margins = src_ws.page_margins
        ws.print_options = src_ws.print_options
        try:
            ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
            ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
            ws.sheet_view.showRowColHeaders = src_ws.sheet_view.showRowColHeaders
            ws.sheet_view.rightToLeft = src_ws.sheet_view.rightToLeft
            ws.sheet_view.tabSelected = src_ws.sheet_view.tabSelected
            ws.sheet_view.topLeftCell = src_ws.sheet_view.topLeftCell
            ws.sheet_view.selection = src_ws.sheet_view.selection
        except Exception:
            pass
    output_wb.save(OUTPUT_FILE)

def process_dispersion_sheet() -> None:
    """分散化シートの行複製・データ挿入・色付け・数式調整"""
    output_wb = load_workbook(OUTPUT_FILE)
    sheet_name = f'{YYMM}{SHEET_DISPERSION}'
    if sheet_name not in output_wb.sheetnames:
        return
    ws = output_wb[sheet_name]
    if SHEET_COMPANY not in output_wb.sheetnames:
        output_wb.save(OUTPUT_FILE)
        return
    df_company = pd.read_excel(OUTPUT_FILE, sheet_name=SHEET_COMPANY, engine='openpyxl')
    data_count = len(df_company)
    template_row = [ws.cell(row=2, column=col) for col in range(1, ws.max_column + 1)]
    sum_template_row = [ws.cell(row=3, column=col) for col in range(1, ws.max_column + 1)]
    ws.delete_rows(3, ws.max_row - 2)
    if data_count > 1:
        ws.insert_rows(3, data_count - 1)
    for i in range(1, data_count):
        for col in range(1, ws.max_column + 1):
            copy_cell(template_row[col-1], ws.cell(row=2 + i, column=col))
    sum_row_idx = 2 + data_count
    for col, sum_tmpl_cell in enumerate(sum_template_row, 1):
        copy_cell(sum_tmpl_cell, ws.cell(row=sum_row_idx, column=col))
    
    kei_col: Optional[int] = None
    sagi_col: Optional[int] = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and '計' in str(cell_value):
            kei_col = col
        elif cell_value and '差額' in str(cell_value):
            sagi_col = col
    
    if kei_col is None:
        kei_col = KEI_COL
    if sagi_col is None:
        sagi_col = SAGI_COL
    
    kei_prev_col = kei_col - 1
    kei_prev_letter = get_column_letter(kei_prev_col)
    kei_letter = get_column_letter(kei_col)
    sagi_letter = get_column_letter(sagi_col)
    
    for i, (idx, row) in enumerate(df_company.iterrows(), 0):
        target_row = 2 + i
        ws.cell(row=target_row, column=1, value=str(row[COL_CUSTOMER_CODE]))
        ws.cell(row=target_row, column=3, value=str(row[COL_CUSTOMER_NAME]))
        ws.cell(row=target_row, column=4, value=float(row[COL_SALES_AMOUNT]))
        ws.cell(row=target_row, column=kei_col, value=f'=SUM(E{target_row}:{kei_prev_letter}{target_row})')
        ws.cell(row=target_row, column=sagi_col, value=f'=D{target_row}-{kei_letter}{target_row}')
    
    for col in range(4, kei_col):
        col_letter = get_column_letter(int(col))
        ws.cell(row=sum_row_idx, column=col, value=f'=SUM({col_letter}2:{col_letter}{1 + data_count})')
    ws.cell(row=sum_row_idx, column=kei_col, value=f'=SUM(E{sum_row_idx}:{kei_prev_letter}{sum_row_idx})')
    ws.cell(row=sum_row_idx, column=sagi_col, value=f'=D{sum_row_idx}-{kei_letter}{sum_row_idx}')
    
    light_green = PatternFill(start_color=COLOR_GROUP, end_color=COLOR_GROUP, fill_type='solid')
    for col in range(4, kei_col):
        ws.cell(row=1, column=col).fill = light_green
    light_yellow = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    ws.cell(row=1, column=kei_col).fill = light_yellow
    
    current_group = None
    color_index = 0
    first_group_processed = False
    for i, (idx, row) in enumerate(df_company.iterrows(), 0):
        target_row = 2 + i
        group = row[COL_GROUP_NAME]
        
        if group != current_group:
            current_group = group
            if first_group_processed:
                color_index = (color_index + 1) % 2
            else:
                first_group_processed = True
        cell_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') if color_index == 0 else PatternFill(start_color=COLOR_GROUP, end_color=COLOR_GROUP, fill_type='solid')
        ws.cell(row=target_row, column=1).fill = cell_fill
        ws.cell(row=target_row, column=3).fill = cell_fill
        b_cell_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') if target_row % 2 == 0 else PatternFill(start_color=COLOR_GROUP, end_color=COLOR_GROUP, fill_type='solid')
        ws.cell(row=target_row, column=2).fill = b_cell_fill
    
    ws['D1'].value = f'{YYMM}{ws["D1"].value if ws["D1"].value else ""}'
    output_wb.save(OUTPUT_FILE)

def copy_sales_db_sheets() -> None:
    """売上DB.xlsxから指定シートをoutput/グループ売上.xlsxに完全コピー"""
    db_wb = load_workbook(SALES_DB_FILE)
    out_wb = load_workbook(OUTPUT_FILE)
    for sheet_name in SALES_DB_SHEETS:
        if sheet_name not in db_wb.sheetnames:
            continue
        if sheet_name in out_wb.sheetnames:
            del out_wb[sheet_name]
        src_ws = db_wb[sheet_name]
        ws = out_wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
        for merged_cell_range in src_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_cell_range))
        for row_idx, row_dim in src_ws.row_dimensions.items():
            ws.row_dimensions[row_idx].height = row_dim.height
        for col, col_dim in src_ws.column_dimensions.items():
            ws.column_dimensions[col].width = col_dim.width
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                copy_cell(cell, new_cell)
                if cell.data_type == 'f':
                    new_cell.value = f'={cell.value}' if not str(cell.value).startswith('=') else cell.value
        if src_ws.auto_filter.ref:
            ws.auto_filter.ref = src_ws.auto_filter.ref
        if src_ws.freeze_panes:
            ws.freeze_panes = src_ws.freeze_panes
        ws.page_setup = src_ws.page_setup
        ws.page_margins = src_ws.page_margins
        ws.print_options = src_ws.print_options
        try:
            ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
            ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
            ws.sheet_view.showRowColHeaders = src_ws.sheet_view.showRowColHeaders
            ws.sheet_view.rightToLeft = src_ws.sheet_view.rightToLeft
            ws.sheet_view.tabSelected = src_ws.sheet_view.tabSelected
            ws.sheet_view.topLeftCell = src_ws.sheet_view.topLeftCell
            ws.sheet_view.selection = src_ws.sheet_view.selection
        except Exception:
            pass
    out_wb.save(OUTPUT_FILE)

def format_monthly_sales_sheet() -> None:
    """当月売上シートの項目名・レイアウトを整理し、順位・得意先・売上を並べる"""
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[YYMM]
    df = pd.read_excel(OUTPUT_FILE, sheet_name=YYMM, engine='openpyxl')
    total_row = df[df.iloc[:, 0] == COL_TOTAL_LABEL]
    data_df = df[df.iloc[:, 0] != COL_TOTAL_LABEL].copy()
    data_df['順位'] = range(1, len(data_df) + 1)
    data_df = data_df[['順位', data_df.columns[0], data_df.columns[1]]]
    ws['A1'] = f'{YYMM}月分'
    ws['A1'].alignment = Alignment(horizontal='left')
    ws['B1'] = ''
    ws['A2'] = LABEL_RANK
    ws['B2'] = LABEL_CUSTOMER
    ws['C2'] = LABEL_CURRENT_SALES
    for i, row in enumerate(data_df.itertuples(index=False), 0):
        ws.cell(row=3 + i, column=1, value=int(row[0]))
        ws.cell(row=3 + i, column=2, value=str(row[1]))
        ws.cell(row=3 + i, column=3, value=float(row[2]))
    if not total_row.empty:
        last_row = 3 + len(data_df)
        ws.cell(row=last_row, column=2, value=LABEL_TOTAL)
        ws.cell(row=last_row, column=3, value=total_row.iloc[0, -1])
    wb.save(OUTPUT_FILE)

def reorder_output_sheets() -> None:
    """アウトプットブックのシート順を指定順に並べ替える"""
    wb = load_workbook(OUTPUT_FILE)
    desired_order = [
        SHEET_CONCAT,
        SHEET_COMPANY,
        f'{YYMM}{SHEET_DISPERSION}',
        PREV_YYMM,
        LAST_MONTH_YYMM,
        YYMM,
        '売上比較',
        f'{YYMM}{SHEET_GROUP_RANK}'
    ]
    ordered_sheets = [wb[s] for s in desired_order if s in wb.sheetnames]
    all_sheets = list(wb.worksheets)
    for s in all_sheets:
        if s not in ordered_sheets:
            ordered_sheets.append(s)
    wb._sheets = ordered_sheets
    wb.save(OUTPUT_FILE)

def create_sales_comparison_sheet_v2() -> None:
    """当月売上・前年・先月の得意先名を正規化して90%以上の類似度でマッチさせ、比較シートを作成"""
    wb = load_workbook(OUTPUT_FILE)
    df_now = pd.read_excel(OUTPUT_FILE, sheet_name=YYMM, engine='openpyxl')
    try:
        df_prev = pd.read_excel(OUTPUT_FILE, sheet_name=PREV_YYMM, engine='openpyxl')
    except Exception:
        print('前年分のデータが見つかりませんでした。')
        return
    try:
        df_last = pd.read_excel(OUTPUT_FILE, sheet_name=LAST_MONTH_YYMM, engine='openpyxl')
    except Exception:
        print('先月分のデータが見つかりませんでした。')
        return
    
    mask_now = df_now.iloc[:, 0].apply(lambda x: str(x).isdigit())
    df_now_data = df_now[mask_now]
    now_ranks = df_now_data.iloc[:, 0].tolist()
    now_names = df_now_data.iloc[:, 1].tolist()
    now_sales = df_now_data.iloc[:, 2].tolist()
    
    mask_prev = df_prev.iloc[:, 0].apply(lambda x: str(x).isdigit())
    df_prev_data = df_prev[mask_prev]
    names_prev = df_prev_data.iloc[:, 1].map(lambda x: str(x).replace(' ', '').replace('　', '') if pd.notnull(x) else '').tolist()
    sales_prev = df_prev_data.iloc[:, 2].tolist()
    orig_names_prev = df_prev_data.iloc[:, 1].tolist()
    
    mask_last = df_last.iloc[:, 0].apply(lambda x: str(x).isdigit())
    df_last_data = df_last[mask_last]
    names_last = df_last_data.iloc[:, 1].map(lambda x: str(x).replace(' ', '').replace('　', '') if pd.notnull(x) else '').tolist()
    sales_last = df_last_data.iloc[:, 2].tolist()
    orig_names_last = df_last_data.iloc[:, 1].tolist()
    
    from collections import defaultdict
    map_prev: Dict[str, float] = defaultdict(float)
    for n, s in zip(names_prev, sales_prev):
        try:
            map_prev[n] += float(s)
        except Exception:
            pass
    map_last: Dict[str, float] = defaultdict(float)
    for n, s in zip(names_last, sales_last):
        try:
            map_last[n] += float(s)
        except Exception:
            pass
    
    def similarity_ratio(str1: str, str2: str) -> float:
        if not str1 or not str2:
            return 0
        str1, str2 = str1.lower(), str2.lower()
        if len(str1) < len(str2):
            str1, str2 = str2, str1
        if len(str2) == 0:
            return 0
        distances = range(len(str2) + 1)
        for i, c1 in enumerate(str1):
            distances_ = [i + 1]
            for j, c2 in enumerate(str2):
                if c1 == c2:
                    distances_.append(distances[j])
                else:
                    distances_.append(1 + min((distances[j], distances[j + 1], distances_[-1])))
            distances = distances_
        return (1 - distances[-1] / len(str1)) * 100
    
    def find_best_match(target_name: str, candidate_names: List[str], threshold: float = SIMILARITY_THRESHOLD) -> Tuple[Optional[str], float]:
        best_match = None
        best_score = 0
        for name in candidate_names:
            score = similarity_ratio(target_name, name)
            if score > best_score and score >= threshold:
                best_score = score
                best_match = name
        return best_match, best_score
    
    rows: List[List[Any]] = []
    matched_prev: Dict[str, float] = defaultdict(float)
    matched_last: Dict[str, float] = defaultdict(float)
    now_normalized = [str(name).replace(' ', '').replace('　', '') if pd.notnull(name) else '' for name in now_names]
    
    for rank, name, sale, norm_name in zip(now_ranks, now_names, now_sales, now_normalized):
        best_prev, score_prev = find_best_match(norm_name, names_prev, SIMILARITY_THRESHOLD)
        c = map_prev[best_prev] if best_prev else None
        if best_prev:
            matched_prev[best_prev] += map_prev[best_prev]
        best_last, score_last = find_best_match(norm_name, names_last, SIMILARITY_THRESHOLD)
        d = map_last[best_last] if best_last else None
        if best_last:
            matched_last[best_last] += map_last[best_last]
        rows.append([rank, name, c, None, d, None, sale])
    
    unmatched_last = [(n, map_last[n], orig_names_last[i]) for i, n in enumerate(names_last) if matched_last[n] == 0]
    unmatched_prev = [(n, map_prev[n], orig_names_prev[i]) for i, n in enumerate(names_prev) if matched_prev[n] == 0]
    unmatched_last_names = [n for n, _, _ in unmatched_last]
    unmatched_prev_names = [n for n, _, _ in unmatched_prev]
    unmatched_last.sort(key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
    unmatched_prev.sort(key=lambda x: x[1] if x[1] is not None else 0, reverse=True)
    
    added_names = set(now_normalized)
    used_prev = set()
    
    for n_last, s_last, orig_name_last in unmatched_last:
        if n_last not in added_names:
            best_prev, score_prev = find_best_match(n_last, unmatched_prev_names, SIMILARITY_THRESHOLD)
            c = None
            if best_prev:
                idx_prev = unmatched_prev_names.index(best_prev)
                c = unmatched_prev[idx_prev][1]
                used_prev.add(best_prev)
            rows.append([None, orig_name_last, c, None, s_last, None, None])
            added_names.add(n_last)
    
    for n_prev, s_prev, orig_name_prev in unmatched_prev:
        if n_prev not in added_names and n_prev not in used_prev:
            best_last, score_last = find_best_match(n_prev, unmatched_last_names, SIMILARITY_THRESHOLD)
            d = None
            if best_last:
                idx_last = unmatched_last_names.index(best_last)
                d = unmatched_last[idx_last][1]
            rows.append([None, orig_name_prev, s_prev, None, d, None, None])
            added_names.add(n_prev)
    
    if '売上比較' in wb.sheetnames:
        del wb['売上比較']
    ws = wb.create_sheet('売上比較')
    ws['A1'] = LABEL_RANK
    ws['B1'] = LABEL_CUSTOMER
    ws['C1'] = LABEL_PREV_SALES
    ws['D1'] = ''
    ws['E1'] = LABEL_LAST_SALES
    ws['F1'] = ''
    ws['G1'] = LABEL_NOW_SALES
    
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 2 and row[0] is None:
                cell.fill = PatternFill(start_color=UNMATCHED_ROW_COLOR, end_color=UNMATCHED_ROW_COLOR, fill_type='solid')
    
    last_row = ws.max_row
    for i in range(2, last_row+1):
        ws.cell(row=i, column=1, value=i-1)
    
    wb.save(OUTPUT_FILE)

def fill_ratio_sheet_from_comparison() -> None:
    """グループ順位シートの4行目（書式ごと）を売上比較のデータ件数分だけ複製し、その上にデータを上書きし、D,F,H,I,J列に数式を自動でセットする"""
    wb = load_workbook(OUTPUT_FILE)
    sheet_name = f'{YYMM}{SHEET_GROUP_RANK}'
    if sheet_name not in wb.sheetnames:
        logger.error(f"グループ順位シートが見つかりません: {sheet_name}")
        return
    ws_ratio = wb[sheet_name]
    if '売上比較' not in wb.sheetnames:
        logger.error('売上比較シートが見つかりません。')
        return
    ws_comp = wb['売上比較']
    comp_data: List[List[Any]] = []
    for row in ws_comp.iter_rows(min_row=2, values_only=True):
        comp_data.append(list(row))
    n = len(comp_data)
    if n == 0:
        logger.warning('売上比較シートにデータがありません。')
        return
    template_row = [ws_ratio.cell(row=4, column=col) for col in range(1, ws_ratio.max_column+1)]
    if n > 1:
        ws_ratio.insert_rows(5, n-1)
        for i in range(n-1):
            for col, tmpl_cell in enumerate(template_row, 1):
                copy_cell(tmpl_cell, ws_ratio.cell(row=5+i, column=col))
    for i, data_row in enumerate(comp_data):
        for j, val in enumerate(data_row, 1):
            ws_ratio.cell(row=4+i, column=j, value=val)
    total_row = 4 + n
    for i in range(n):
        row_idx = 4 + i
        # D,F,H,I,J列の数式を必ずセット
        ws_ratio.cell(row=row_idx, column=4).value = f"=C{row_idx}/C{total_row}"
        ws_ratio.cell(row=row_idx, column=6).value = f"=E{row_idx}/E{total_row}"
        ws_ratio.cell(row=row_idx, column=8).value = f"=G{row_idx}/G{total_row}"
        ws_ratio.cell(row=row_idx, column=9).value = f'=IF(OR(C{row_idx}="",C{row_idx}=0),"",G{row_idx}/C{row_idx})'
        ws_ratio.cell(row=row_idx, column=10).value = f'=IF(OR(E{row_idx}="",E{row_idx}=0),"",G{row_idx}/E{row_idx})'
    sum_start = 4
    sum_end = total_row - 1
    for col in range(3, 9):
        col_letter = get_column_letter(col)
        ws_ratio.cell(row=total_row, column=col).value = f"=SUM({col_letter}{sum_start}:{col_letter}{sum_end})"
    ws_ratio.cell(row=total_row, column=4).value = f"=SUM(D{sum_start}:D{sum_end})"
    ws_ratio.cell(row=total_row, column=9).value = f'=IF(OR(C{total_row}="",C{total_row}=0),"",G{total_row}/C{total_row})'
    ws_ratio.cell(row=total_row, column=10).value = f'=IF(OR(E{total_row}="",E{total_row}=0),"",G{total_row}/E{total_row})'
    yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    light_green = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
    light_yellow = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    for col in range(1, 12):
        ws_ratio.cell(row=1, column=col).fill = yellow
    for row in range(2, 4):
        for col in range(1, 12):
            ws_ratio.cell(row=row, column=col).fill = light_green
    for col in range(1, 12):
        ws_ratio.cell(row=total_row, column=col).fill = light_yellow
    ws_ratio['A1'].value = f'{YYMM}{ws_ratio["A1"].value if ws_ratio["A1"].value else ""}'
    ws_ratio['G3'].value = f'{YYMM}{ws_ratio["G3"].value if ws_ratio["G3"].value else ""}'
    ws_ratio['C3'].value = f'{PREV_YYMM}{ws_ratio["C3"].value if ws_ratio["C3"].value else ""}'
    ws_ratio['E3'].value = f'{LAST_MONTH_YYMM}{ws_ratio["E3"].value if ws_ratio["E3"].value else ""}'
    wb.save(OUTPUT_FILE)

def format_monthly_sales_borders() -> None:
    """当月売上シートのA〜C列で合計行まで格子罫線を付け、ボールドを標準フォントに戻す"""
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[YYMM]
    max_row = ws.max_row
    last_row = 1
    for row in range(1, max_row+1):
        if ws.cell(row=row, column=2).value == '合計':
            last_row = row
            break
        if ws.cell(row=row, column=1).value not in [None, '']:
            last_row = row
    thin = Side(border_style=BORDER_STYLE, color=BORDER_COLOR)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in range(1, last_row+1):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if cell.font.bold:
                cell.font = Font(name=cell.font.name, size=cell.font.size, bold=False, italic=cell.font.italic, color=cell.font.color)
    wb.save(OUTPUT_FILE)

def optimize_all_sheet_column_widths() -> None:
    """全シート・全列・全行を対象に、セル内容に合わせて列幅を最適化する"""
    wb = load_workbook(OUTPUT_FILE)
    for ws in wb.worksheets:
        if ws.title in EXCLUDE_AUTO_WIDTH:
            continue
        for col in ws.columns:
            max_length = 0
            col_idx = col[0].column
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    width = sum(2 if ord(c) > 127 else 1 for c in val.strip())
                    if width > max_length:
                        max_length = width
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(OUTPUT_FILE)

def write_all_sheets_with_pandas(all_df, group_sum):
    """pandasで全シートを書き出す（連結・グループ合計・会社別）"""
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
        all_df.to_excel(writer, index=False, sheet_name=SHEET_CONCAT)
        group_sum.to_excel(writer, index=False, sheet_name=YYMM)
        all_df_sorted = all_df.copy()
        group_order = group_sum[group_sum['グループ名'] != '合計']['グループ名'].tolist()
        group_rank = {g: i for i, g in enumerate(group_order)}
        all_df_sorted['グループ売上順位'] = all_df_sorted[COL_GROUP_NAME].map(group_rank).fillna(len(group_order))
        all_df_sorted = all_df_sorted.sort_values(['グループ売上順位', COL_SALES_AMOUNT], ascending=[True, False])
        all_df_sorted[COL_SALES_AMOUNT] = pd.to_numeric(all_df_sorted[COL_SALES_AMOUNT], errors='coerce')
        all_df_sorted.drop(columns=['グループ売上順位']).to_excel(writer, index=False, sheet_name=SHEET_COMPANY)
    # 連結シート色付け
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_CONCAT]
    format_concat_sheet_colors(ws, all_df)
    wb.save(OUTPUT_FILE)

def main() -> None:
    """全体の処理フローを統括"""
    setup_logging()
    logger.info("=== 売上データ自動処理スクリプト開始 ===")
    try:
        global YYMM, PREV_YYMM, LAST_MONTH_YYMM, TEMPLATE_SHEETS, SALES_DB_SHEETS, OUTPUT_FILE
        YYMM = input('何年何月分のデータですか？(例:2025年05月→2505): ').strip()
        if not YYMM.isdigit() or len(YYMM) != 4:
            logger.error(f"不正なYYMM形式: {YYMM}")
            print('正しいYYMM形式で入力してください (例:2505)')
            return
        PREV_YYMM = get_prev_yymm(YYMM)
        LAST_MONTH_YYMM = get_last_month_yymm(YYMM)
        TEMPLATE_SHEETS = [f'{YYMM}{SHEET_DISPERSION}', f'{YYMM}{SHEET_GROUP_RANK}']
        SALES_DB_SHEETS = [PREV_YYMM, LAST_MONTH_YYMM]
        # ここでOUTPUT_FILEの先頭にYYMMを付与
        import os
        base_dir = os.path.dirname(config['output']['file'])
        base_name = os.path.basename(config['output']['file'])
        OUTPUT_FILE = os.path.join(base_dir, f"{YYMM}_{base_name}") if base_dir else f"{YYMM}_{base_name}"
        logger.info(f"処理対象: YYMM={YYMM} 前年={PREV_YYMM} 先月={LAST_MONTH_YYMM}")
        print(f"[INFO] YYMM={YYMM} 前年={PREV_YYMM} 先月={LAST_MONTH_YYMM} で処理を開始します")
        logger.info("データ前処理開始")
        all_list = load_and_preprocess_data()
        if not all_list:
            logger.error("有効なデータがありません")
            print('[ERROR] 有効なデータがありません')
            return
        print("[OK] データ前処理 完了")
        logger.info("マスタマージ開始")
        all_df = pd.concat(all_list, ignore_index=True)
        all_df = merge_with_master(all_df)
        print("[OK] マスタマージ・整形 完了")
        logger.info("グループ合計計算開始")
        group_sum = all_df.groupby(COL_GROUP_NAME, dropna=False)[COL_SALES_AMOUNT].sum().reset_index()
        group_sum = group_sum.rename(columns={COL_GROUP_NAME: COL_GROUP_LABEL, COL_SALES_AMOUNT: COL_TOTAL_SALES})
        group_sum = group_sum.sort_values(COL_TOTAL_SALES, ascending=False)
        total = group_sum[COL_TOTAL_SALES].sum()
        total_row = pd.DataFrame({COL_GROUP_LABEL: [COL_TOTAL_LABEL], COL_TOTAL_SALES: [total]})
        group_sum = pd.concat([group_sum, total_row], ignore_index=True)
        group_sum[COL_TOTAL_SALES] = pd.to_numeric(group_sum[COL_TOTAL_SALES], errors='coerce')
        print("[OK] グループ合計出力 完了")
        # pandasで全シート書き出し
        write_all_sheets_with_pandas(all_df, group_sum)
        print("[OK] pandasで全シート書き出し 完了")
        # ここからopenpyxlで加工
        copy_template_sheets()
        process_dispersion_sheet()
        copy_sales_db_sheets()
        format_monthly_sales_sheet()
        create_sales_comparison_sheet_v2()
        fill_ratio_sheet_from_comparison()
        format_monthly_sales_borders()
        optimize_all_sheet_column_widths()
        reorder_output_sheets()
        print("[END] 全処理正常終了！")
        logger.info("=== 全処理正常終了 ===")
    except SalesDataError as e:
        logger.error(f"売上データ処理エラー: {e}")
        print(f"[ERROR] 処理エラー: {e}")
    except KeyboardInterrupt:
        logger.info("ユーザーによる処理中断")
        print("\n[INFO] 処理が中断されました")
    except Exception as e:
        logger.error(f"予期しないエラー: {e}", exc_info=True)
        print(f"[ERROR] 予期しないエラーが発生しました: {e}")
    finally:
        logger.info("=== 処理終了 ===")

if __name__ == '__main__':
    main()
